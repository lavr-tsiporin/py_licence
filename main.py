import eel
import requests
import openpyxl
import datetime
import os
from dotenv import load_dotenv

dotenv_path = os.path.join(os.path.dirname(__file__), '.env')
if os.path.exists(dotenv_path):
    load_dotenv(dotenv_path)
   
eel.init('web')

@eel.expose
def transfer_data(nameFile, dataFile):
    nowDate = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    name = nameFile.split('.')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'car_number'
    ws['B1'] = 'license_status'
    ws['C1'] = 'license_change_date'
    ws['D1'] = 'license_change_date'

    arrayNumberMachine = dataFile.replace('\r', '').split('\n') 
    arrayNumberMachineNoDooble = list(filter(None, list(set(arrayNumberMachine))))
    eel.sizeData(len(arrayNumberMachineNoDooble), len(arrayNumberMachine))
    eel.status('Progress')

    for i in arrayNumberMachineNoDooble:
        eel.progress(str(arrayNumberMachineNoDooble.index(i)+1))
        dataResult = requestData(i)
        index = arrayNumberMachineNoDooble.index(i)+2
        ws[f'A{index}'] = i
        ws[f'B{index}'] = dataResult[1]
        ws[f'C{index}'] = dataResult[2]
        ws[f'D{index}'] = dataResult[3]
    wb.save(f'./{name[0]}_{str(nowDate)}.xlsx')
    eel.status(f'Finish, file created {name[0]}_{str(nowDate)}')

def requestData(number):
    try:
        urlFetch = f'{os.getenv("URL_FETCH_ONE")}={number}&{os.getenv("URL_FETCH_TWO")}'
        urlCookie = os.getenv("URL_MAIN")

        r = requests.get(urlCookie)
    
        cookie = r.cookies
        fetchData = requests.get(urlFetch, headers={'cookie': os.getenv("ID")}, cookies=cookie)

        data = fetchData.json()

        if int(data['Count']) == 0:
            cond = ''
            date1 = ''
            date2 = ''
        else:
            all = data['Infos']
            dataCondition = [item for item in all if item['Condition'] == 'Действующее']
            if len(dataCondition) >= 1:
                last = dataCondition[-1]
                cond = last['Condition']
                date = last['ValidityDate'].split(' ')
                date1 = date[1]
                date2 = date[-1]
            else:
                last = data['Infos'][-1]
                cond = last['Condition']
                date = last['ValidityDate'].split(' ')
                date1 = date[1]
                date2 = date[-1]
            
        result = [number, cond, date1, date2]
        return result
    except:
        return [f'ошибка {number}', '', '', '']

eel.start('main.html', size=(600, 400))